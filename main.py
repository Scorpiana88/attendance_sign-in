import openpyxl
import keyboard

def create_excel_file():
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define column names in the Excel sheet
    column_names = ["Month", "Year", "Youth Barcode ID", "Student Name", "Grade Level",
                    "Did not attend this month?", "Reading Level/Literacy Support",
                    "Reading Level 3 Hours", "Math Support Level", "Math Level 3 Hours",
                    "Sports", "STEM", "Arts", "Leadership"]

    # Add day columns (1 to 31)
    for day in range(1, 32):
        column_names.append(f"Day {day}")

    # Write column names to the first row
    for col_num, column_name in enumerate(column_names, start=1):
        ws.cell(row=1, column=col_num, value=column_name)

    # Save the workbook
    wb.save("student_data.xlsx")

def scan_student_data():
    # Load existing Excel file or create a new one if not exists
    try:
        wb = openpyxl.load_workbook("student_data.xlsx")
        ws = wb.active
    except FileNotFoundError:
        create_excel_file()
        wb = openpyxl.load_workbook("student_data.xlsx")
        ws = wb.active

    print("Excel file loaded. Start scanning...")

    # Listen for barcode scanner input
    while True:
        barcode = keyboard.read_event().name
        if barcode == "esc":  # Exit if 'esc' key is pressed
            break

        # Check if the student data already exists
        student_found = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3, values_only=True):
            if barcode == row[0]:
                print("Student already exists.")
                student_found = True
                break

        if not student_found:
            # Get input for each column
            month = input("Enter Month (e.g., January): ")
            year = input("Enter Year: ")
            student_name = input("Enter Student Name: ")
            grade_level = input("Enter Grade Level: ")
            attendance = "Yes"  # Default value
            reading_level_support = "Level 1"  # Default value
            reading_level_3_hours = input("Reading Level 3 Hours: ")
            math_support_level = "Level 1"  # Default value
            math_level_3_hours = input("Math Level 3 Hours: ")
            sports = "No"  # Default value
            stem = "No"  # Default value
            arts = "No"  # Default value
            leadership = "No"  # Default value

            # Append the data to the Excel file
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=month)
            ws.cell(row=next_row, column=2, value=year)
            ws.cell(row=next_row, column=3, value=barcode)
            ws.cell(row=next_row, column=4, value=student_name)
            ws.cell(row=next_row, column=5, value=grade_level)
            ws.cell(row=next_row, column=6, value=attendance)
            ws.cell(row=next_row, column=7, value=reading_level_support)
            ws.cell(row=next_row, column=8, value=reading_level_3_hours)
            ws.cell(row=next_row, column=9, value=math_support_level)
            ws.cell(row=next_row, column=10, value=math_level_3_hours)
            ws.cell(row=next_row, column=11, value=sports)
            ws.cell(row=next_row, column=12, value=stem)
            ws.cell(row=next_row, column=13, value=arts)
            ws.cell(row=next_row, column=14, value=leadership)

            # Set default values for day columns
            for day in range(15, 46):
                ws.cell(row=next_row, column=day, value="N/A")

            # Save the workbook
            wb.save("student_data.xlsx")

            print("Student scanned and data saved.")

    print("Exiting...")

if __name__ == "__main__":
    scan_student_data()



